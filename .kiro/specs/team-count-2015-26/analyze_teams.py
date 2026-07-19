#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
シーズン別 ユニーク参加チーム数 集計・分析スクリプト（参照非依存）。

入力: raw_team_by_season.tsv  (列: season_year<TAB>count<TAB>team_name)
       ※ entry基準(entry_status=0, deleted=0, racers結合あり)・15-16〜25-26 で
         (season_year, team_name) ごとに COUNT(*) を集計した生データ。
         NULL は '<<NULL>>' というセンチネルで表現。生成SQLは test-results.md / Runbook 参照。

モード:
  --preview   レビューゲート①の材料を生成・表示する（集計確定前）
              - 無所属として除外する値の一覧と件数
              - 正規化(NFKC+大小文字+空白除去)で統合される生チーム名グループ一覧
              - 見落とし検知用のボーダーライン候補
  --build     レビュー承認後に最終集計を行い Excel を生成する

正規化方針（決定#4=積極正規化）:
  norm_key = NFKC正規化 → 全空白(半角/全角/タブ等)除去 → casefold(大小文字統一)
  これにより「ＡＢＣ」「ABC」「a b c」「ＡＢＣ 」等を同一チームとみなす。

無所属判定（決定#3）:
  norm_key が NOTEAM_KEYS のいずれかに一致する行を「無所属」として
  ユニークチーム数から除外し、別途エントリー件数として集計する。
"""
import argparse
import os
import re
import sys
import unicodedata
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
TSV = os.path.join(HERE, "raw_team_by_season.tsv")
NULL_SENTINEL = "<<NULL>>"

# シーズン開始年(int) -> 表示ラベル
def season_label(y):
    return f"{y % 100:02d}-{(y + 1) % 100:02d}"

SEASON_YEARS = list(range(2015, 2026))  # 2015(15-16) .. 2025(25-26)

# --- 無所属とみなす正規化キー（決定#3, レビュー対象） -----------------------
# norm() を通した後の値で比較する。空文字・各種ダッシュ・「なし」系・NULL。
NOTEAM_RAWS = [
    "", "<<NULL>>",
    "-", "−", "―", "ー", "‐", "–", "—",          # 各種ダッシュ/長音記号
    "なし", "無し", "無", "無所属", "所属なし", "ナシ",
    ".", "．", "・", "/",
]


def norm(s):
    """正規化キーを返す。"""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"\s+", "", s)       # 全角含む全空白除去
    return s.casefold()


NOTEAM_KEYS = {norm(x) for x in NOTEAM_RAWS if x != NULL_SENTINEL}
NOTEAM_KEYS.add(NULL_SENTINEL.casefold())  # NULL センチネルはそのまま無所属扱い

# --- レビューゲート①で追加確定した無所属ルール（材料C 4バケツ） -------------
# 完全一致で無所属扱いにする norm_key（「フリー」は部分一致しない＝実在チーム保護）
NOTEAM_EXACT = {norm(x) for x in ["0", "フリー", "個人参加"]}
# 部分文字列で無所属扱い（材料Cで全件が無所属と確認済み）
NOTEAM_SUBSTR = ["なし", "無所属", "未定", "個人"]


def is_noteam(raw):
    if raw == NULL_SENTINEL:
        return True
    disp = raw.strip()
    k = norm(raw)
    if k in NOTEAM_KEYS:
        return True
    # 材料C-1 プレースホルダ0 / 材料C-3 個人参加・フリー 等（完全一致）
    if k in NOTEAM_EXACT:
        return True
    # 材料C-2 記号のみ（CJK/英数を含まない）
    if k and re.fullmatch(r"[\W_]+", k):
        return True
    # 材料C-4 1文字の漢字/英字（無所属プレースホルダとして除外）
    if len(disp) <= 1:
        return True
    # 材料C-3 明示的な「無所属/なし/未定/個人」句（部分一致）
    if any(t in disp for t in NOTEAM_SUBSTR):
        return True
    return False


def load_rows():
    rows = []
    with open(TSV, encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            parts = line.split("\t")
            if len(parts) < 3:
                # team_name が空文字のケース: 末尾フィールドが空
                parts = (parts + ["", "", ""])[:3]
            year = int(parts[0])
            cnt = int(parts[1])
            name = "\t".join(parts[2:])  # 念のため
            rows.append((year, cnt, name))
    return rows


def aggregate(rows):
    """
    戻り値:
      season_teams[year] -> dict(norm_key -> {'count':int, 'raws':set, 'rep':str})
      noteam[year] -> エントリー件数
    rep(代表表示名)= そのキー内で最も件数の多い生名称。
    """
    season_teams = defaultdict(lambda: defaultdict(lambda: {"count": 0, "raws": defaultdict(int)}))
    noteam = defaultdict(int)
    for year, cnt, raw in rows:
        if is_noteam(raw):
            noteam[year] += cnt
            continue
        k = norm(raw)
        disp = raw.strip()
        d = season_teams[year][k]
        d["count"] += cnt
        d["raws"][disp] += cnt
    return season_teams, noteam


def rep_name(raws):
    return max(raws.items(), key=lambda kv: (kv[1], kv[0]))[0]


def cmd_preview(rows):
    # --- 無所属の集計 ---
    noteam_detail = defaultdict(int)
    for year, cnt, raw in rows:
        if is_noteam(raw):
            key = "<<NULL>>" if raw == NULL_SENTINEL else (raw.strip() or "(空文字)")
            noteam_detail[key] += cnt

    print("=" * 70)
    print("【レビュー材料 A】無所属として除外する値（決定#3）")
    print("=" * 70)
    total_noteam = 0
    for k, c in sorted(noteam_detail.items(), key=lambda x: -x[1]):
        print(f"  {c:>7,}  {k!r}")
        total_noteam += c
    print(f"  --- 無所属 合計エントリー件数: {total_noteam:,}")

    # --- 正規化で統合される生名称グループ ---
    key_to_raws = defaultdict(lambda: defaultdict(int))
    for year, cnt, raw in rows:
        if is_noteam(raw):
            continue
        key_to_raws[norm(raw)][raw.strip()] += cnt

    merged = {k: v for k, v in key_to_raws.items() if len(v) >= 2}
    print()
    print("=" * 70)
    print("【レビュー材料 B】正規化で“同一チーム”に統合される生名称グループ")
    print("  （表記ゆれが本当に同一チームか、誤統合がないか確認してください）")
    print("=" * 70)
    print(f"  統合グループ数: {len(merged)}（2つ以上の生名称が1キーに収束）")
    print()
    # 件数合計の多い順に全件
    for k, raws in sorted(merged.items(), key=lambda kv: -sum(kv[1].values())):
        items = sorted(raws.items(), key=lambda x: -x[1])
        rep = items[0][0]
        variants = " | ".join(f"{n!r}×{c}" for n, c in items)
        print(f"  [{rep}]  {variants}")

    # --- ボーダーライン候補（無所属の見落とし検知） ---
    print()
    print("=" * 70)
    print("【レビュー材料 C】無所属の見落とし候補（除外リスト外で“無所属っぽい”値）")
    print("=" * 70)
    seen = defaultdict(int)
    for year, cnt, raw in rows:
        if is_noteam(raw):
            continue
        disp = raw.strip()
        k = norm(raw)
        # 記号のみ / 1文字 / 「なし」「無所属」を含む など
        if (k and (re.fullmatch(r"[\W_]+", k) or len(disp) <= 1
                   or any(t in disp for t in ("なし", "無所属", "未定", "フリー", "個人")))):
            seen[disp] += cnt
    if seen:
        for k, c in sorted(seen.items(), key=lambda x: -x[1]):
            print(f"  {c:>7,}  {k!r}")
    else:
        print("  （該当なし）")

    # --- 暫定サマリ（参考: この設定での各シーズン ユニーク数） ---
    season_teams, noteam = aggregate(rows)
    print()
    print("=" * 70)
    print("【参考】現設定での暫定 シーズン別ユニークチーム数（承認後に確定）")
    print("=" * 70)
    print(f"  {'シーズン':<8}{'ユニーク数':>10}{'無所属件数':>12}")
    prev = None
    for y in SEASON_YEARS:
        u = len(season_teams.get(y, {}))
        nt = noteam.get(y, 0)
        diff = "" if prev is None else f"({u - prev:+d})"
        print(f"  {season_label(y):<8}{u:>10}{nt:>12,}  {diff}")
        prev = u


def build_series(rows):
    """最終集計の SERIES と per-season team lists を返す。"""
    season_teams, noteam = aggregate(rows)
    series = []  # (label, unique_count, noteam_count)
    team_lists = {}  # label -> [ (rep_name, entry_count) ... ] sorted
    for y in SEASON_YEARS:
        teams = season_teams.get(y, {})
        label = season_label(y)
        series.append((label, len(teams), noteam.get(y, 0)))
        lst = []
        for k, d in teams.items():
            lst.append((rep_name(d["raws"]), d["count"]))
        lst.sort(key=lambda x: (-x[1], x[0]))
        team_lists[label] = lst
    return series, team_lists


def cmd_build(rows, out):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    series, team_lists = build_series(rows)
    wb = openpyxl.Workbook()

    # --- シート1: シーズン別ユニーク数 + 前年比 ---
    ws = wb.active
    ws.title = "ユニークチーム数"
    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="DDDDDD")
    heads = ["シーズン", "ユニークチーム数", "前年比", "前年比率", "無所属エントリー件数"]
    for c, h in enumerate(heads, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    prev = None
    for i, (label, u, nt) in enumerate(series):
        r = i + 2
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=u).number_format = "#,##0"
        if prev is not None:
            d = ws.cell(row=r, column=3, value=f"=B{r}-B{r-1}")
            d.number_format = "#,##0;-#,##0"
            rt = ws.cell(row=r, column=4, value=f"=IF(B{r-1}=0,\"\",B{r}/B{r-1}-1)")
            rt.number_format = "0.0%"
        ws.cell(row=r, column=5, value=nt).number_format = "#,##0"
        prev = u
    widths = [10, 18, 10, 10, 20]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    # --- シート2..: 各シーズンの参加チームリスト ---
    for label, lst in team_lists.items():
        s = wb.create_sheet(title=f"{label} チーム")
        for c, h in enumerate(["#", "チーム名", "エントリー件数"], 1):
            cell = s.cell(row=1, column=c, value=h)
            cell.font = bold
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
        for i, (name, cnt) in enumerate(lst, 1):
            s.cell(row=i + 1, column=1, value=i)
            s.cell(row=i + 1, column=2, value=name)
            s.cell(row=i + 1, column=3, value=cnt).number_format = "#,##0"
        s.column_dimensions["A"].width = 6
        s.column_dimensions["B"].width = 40
        s.column_dimensions["C"].width = 14
        s.freeze_panes = "A2"

    wb.save(out)
    print(f"[build] -> {out}")
    for label, u, nt in series:
        print(f"  {label}: unique={u}  noteam={nt}  (list rows={len(team_lists[label])})")
    # 検算: リスト行数 == ユニーク数
    ok = all(len(team_lists[l]) == u for l, u, _ in series)
    print(f"  検算[リスト行数==ユニーク数]: {'OK' if ok else 'NG'}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--preview", action="store_true")
    p.add_argument("--build", action="store_true")
    p.add_argument("--out", default=os.path.join(HERE, "..", "..", "..", "tmp",
                                                  "team_unique_2015-26.xlsx"))
    a = p.parse_args()
    rows = load_rows()
    if a.preview:
        cmd_preview(rows)
    elif a.build:
        os.makedirs(os.path.dirname(os.path.abspath(a.out)), exist_ok=True)
        cmd_build(rows, os.path.abspath(a.out))
    else:
        p.error("--preview か --build を指定してください")


if __name__ == "__main__":
    main()
