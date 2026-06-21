#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
シクロクロス シーズン別 エントリー数・出走数の「通しでカウント」シートを生成する。

2つのモードを持つ:

  (1) テンプレートモード（デフォルト）
      既存ブック（参照ファイル or 前年の出力）を読み込み、最新シーズン行を追記する。
      他シート・書式をすべて温存する。参照ファイルが存在する場合に推奨。
        例: python3 build_xlsx.py --season 25-26

  (2) from-scratch モード（参照ファイル非依存）
      下記 SERIES（このスクリプトに埋め込んだ全シーズンの値）だけを入力に、
      「通しでカウント」シートを 1 枚持つブックをゼロから生成する。
      参照ファイルが失われても再現可能。
        例: python3 build_xlsx.py --from-scratch

値の出所・集計SQLは docs/specs/entry-count-2025-26/test-results.md を参照。
来年以降は SERIES に1行追加し、いずれかのモードを実行する。

シート仕様（test-results.md「出力フォーマット仕様」と一致させること）:
  - 行3: E3 = '前年比'（E,F列の見出し）
  - 行4: C='entry', D='started', E='entry', F='started'
  - 行5以降: B=シーズン, C=entry(#,##0), D=started(#,##0),
             E=前年比entry(=C-前年C, #,##0), F=前年比started(=D-前年D, #,##0)
             ※先頭シーズン行(行5)は前年が無いため E,F は空
  - 列幅: B6.6 / C7.6 / D8.6 / E7.2 / F8.6
"""
from copy import copy
import argparse
import os
import openpyxl
from openpyxl.styles import Font, Alignment

REPO = os.environ.get("CYCLOX_REPO", "/Users/kyamady/workspace/cyclox2_docker")

SHEET = "通しでカウント"
NUMFMT = "#,##0"
COL_WIDTHS = {"B": 6.6, "C": 7.6, "D": 8.6, "E": 7.2, "F": 8.6}
HEADER_ROW3 = {5: "前年比"}                                  # E3
HEADER_ROW4 = {3: "entry", 4: "started", 5: "entry", 6: "started"}  # C4..F4
DATA_START = 5  # 行5から実データ

# --- シーズン別 集計値（単一の真実。来年は1行追加する） -------------------
# (シーズン, entry, started)  ※ test-results.md の検算結果に基づく
SERIES = [
    ("15-16", 18458, 15074),
    ("16-17", 20088, 16411),
    ("17-18", 21266, 17252),
    ("18-19", 21716, 18721),
    ("19-20", 21825, 18387),
    ("20-21", 11973, 10511),
    ("21-22", 18841, 16313),
    ("22-23", 21508, 18962),
    ("23-24", 22052, 19111),
    ("24-25", 21644, 19094),
    ("25-26", 21604, 18802),
]


def series_lookup(season):
    for s, e, st in SERIES:
        if s == season:
            return e, st
    raise SystemExit(f"SERIES に {season} がありません。先にSERIESへ追記してください。")


# --- モード1: 既存ブックに最新行を追記 ------------------------------------
def append_to_template(src, out, season):
    entry, started = series_lookup(season)
    wb = openpyxl.load_workbook(src)  # data_only=False: 書式・数式を温存
    ws = wb[SHEET]

    # B列がシーズン文字列の最終行を探す
    last_row = None
    for r in range(ws.max_row, 1, -1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and "-" in v:
            last_row = r
            break
    if last_row is None:
        raise SystemExit("既存のシーズン行が見つかりません")
    if ws.cell(last_row, 2).value == season:
        raise SystemExit(f"{season} は既に {SHEET} の row={last_row} に存在します")

    new_row = last_row + 1
    values = {
        2: season, 3: entry, 4: started,
        5: f"=C{new_row}-C{last_row}",
        6: f"=D{new_row}-D{last_row}",
    }
    for col in range(2, 7):
        src_cell = ws.cell(row=last_row, column=col)
        dst = ws.cell(row=new_row, column=col)
        dst.value = values[col]
        if src_cell.has_style:
            dst.font = copy(src_cell.font)
            dst.border = copy(src_cell.border)
            dst.fill = copy(src_cell.fill)
            dst.number_format = copy(src_cell.number_format)
            dst.protection = copy(src_cell.protection)
            dst.alignment = copy(src_cell.alignment)

    wb.calculation.fullCalcOnLoad = True  # 他シートの数式をExcel起動時に再計算
    wb.save(out)
    print(f"[template] {src} -> {out}")
    print(f"  追記 row={new_row}: {season} entry={entry} started={started} (前年比は数式)")


# --- モード2: SERIESからシートをゼロ生成（参照ファイル非依存） -------------
def build_from_scratch(out):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET

    for col_letter, w in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = w

    for col, text in HEADER_ROW3.items():
        ws.cell(row=3, column=col, value=text)
    for col, text in HEADER_ROW4.items():
        ws.cell(row=4, column=col, value=text)

    for i, (season, entry, started) in enumerate(SERIES):
        r = DATA_START + i
        ws.cell(row=r, column=2, value=season)
        c = ws.cell(row=r, column=3, value=entry); c.number_format = NUMFMT
        d = ws.cell(row=r, column=4, value=started); d.number_format = NUMFMT
        if i > 0:  # 先頭シーズンは前年が無いので前年比なし
            e = ws.cell(row=r, column=5, value=f"=C{r}-C{r-1}"); e.number_format = NUMFMT
            f = ws.cell(row=r, column=6, value=f"=D{r}-D{r-1}"); f.number_format = NUMFMT

    wb.calculation.fullCalcOnLoad = True
    wb.save(out)
    print(f"[from-scratch] -> {out}")
    print(f"  {len(SERIES)} シーズン (行{DATA_START}..{DATA_START+len(SERIES)-1}) を生成")


def main():
    latest = SERIES[-1][0]
    p = argparse.ArgumentParser(description="通しでカウント シート生成")
    p.add_argument("--from-scratch", action="store_true",
                   help="参照ファイルを使わず SERIES からゼロ生成する")
    p.add_argument("--season", default=latest, help=f"追記するシーズン (既定: {latest})")
    p.add_argument("--src", default=os.path.join(REPO, "tmp", "20250416_entry_racers.xlsx"),
                   help="テンプレートモードの入力ブック（参照 or 前年出力）")
    p.add_argument("--out", default=None, help="出力パス")
    a = p.parse_args()

    if a.from_scratch:
        out = a.out or os.path.join(REPO, "tmp", f"entry_racers_{latest}_from_scratch.xlsx")
        build_from_scratch(out)
    else:
        out = a.out or os.path.join(REPO, "tmp", "20260621_entry_racers.xlsx")
        append_to_template(a.src, out, a.season)


if __name__ == "__main__":
    main()
